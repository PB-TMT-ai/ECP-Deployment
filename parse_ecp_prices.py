"""
parse_ecp_prices.py — Reads the ECP price update Excel and outputs structured JSON.

The Excel (e.g. Book 70.xlsx) has regional pricing blocks side-by-side in Sheet1.
This parser uses exact column mappings discovered from the file structure.

Key rows:
  - Row 20: Discount % in decimal (offset columns, e.g. 0.03 = 3%)
  - Row 28: ECP for one 12mm bar
  - Row 29: ECP for one 10mm bar
  - Row 30: ECP for one 8mm bar
  - Row 31: ECP for one 16mm bar
  - Row 32: ECP for one 20mm bar
  - Row 33: ECP for one 25mm bar

Usage:
  python tools/parse_ecp_prices.py <path_to_excel> <effective_date>

Output: JSON to stdout
"""

import json
import sys
import os

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string as col_idx
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Exact column mappings: region -> (Fe550_price_col, Fe550D_price_col, Fe550_disc_col, Fe550D_disc_col)
REGION_COLUMNS = {
    "Agra":          ("K",  "L",  "M",  "N"),
    "Delhi":         ("AB", "AC", "AD", "AE"),
    "Sonipat":       ("AM", "AN", "AO", "AP"),
    "Himachal":      ("AX", "AY", "AZ", "BA"),
    "Jammu":         ("BC", "BD", "BE", "BF"),
    "Kashmir":       ("BH", "BI", "BJ", "BK"),
    "Rajasthan":     ("BM", "BN", "BO", "BP"),
    "Uttarakhand":   ("BR", "BS", "BT", "BU"),
    "Punjab":        ("BW", "BX", "BY", "BZ"),
    "Bihar":         ("CI", "CJ", "CK", "CL"),
    "Jharkhand":     ("CN", "CO", "CP", "CQ"),
    "West Bengal":   ("CS", "CT", "CU", "CV"),
    "Orissa":        ("CX", "CY", "CZ", "DA"),
    "Madhya Pradesh":("DC", "DD", "DE", "DF"),
    "Chattisgarh":   ("DH", "DI", "DJ", "DK"),
}

# Region -> Canva state name
REGION_TO_STATE = {
    "Agra": "Uttar Pradesh",
    "Delhi": "Delhi",
    "Sonipat": "Haryana",
    "Himachal": "Himachal Pradesh",
    "Jammu": "Jammu",
    "Kashmir": "Kashmir",
    "Rajasthan": "Rajasthan",
    "Uttarakhand": "Uttarakhand",
    "Punjab": "Punjab",
    "Bihar": "Bihar",
    "Jharkhand": "Jharkhand",
    "West Bengal": "West Bengal",
    "Orissa": "Odisha",
    "Madhya Pradesh": "Madhya Pradesh",
    "Chattisgarh": "Chattisgarh",
}

# Row -> bar section (order: 8mm, 10mm, 12mm, 16mm, 20mm, 25mm)
BAR_ROWS = {
    30: "8mm",
    29: "10mm",
    28: "12mm",
    31: "16mm",
    32: "20mm",
    33: "25mm",
}

# Which states appear in which Canva design
FE550_STATES = [
    "Bihar", "Jharkhand", "West Bengal", "Odisha", "Madhya Pradesh",
    "Chattisgarh", "Jammu", "Kashmir", "Uttarakhand", "Himachal Pradesh",
    "Uttar Pradesh", "Delhi", "Haryana", "Rajasthan", "Punjab",
]
FE550D_STATES = [
    "Uttar Pradesh", "Delhi", "Haryana", "Rajasthan", "Punjab",
]


def parse_excel(filepath, effective_date):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    all_prices = {}  # state -> {"Fe 550": {...}, "Fe 550D": {...}}

    for region, (fe550_col, fe550d_col, fe550_disc, fe550d_disc) in REGION_COLUMNS.items():
        state = REGION_TO_STATE[region]
        state_data = {}

        for grade, price_col_str, disc_col_str in [
            ("Fe 550", fe550_col, fe550_disc),
            ("Fe 550D", fe550d_col, fe550d_disc),
        ]:
            price_col = col_idx(price_col_str)
            disc_col = col_idx(disc_col_str)

            # Extract per-bar prices
            prices = {}
            for row_num, section in BAR_ROWS.items():
                val = ws.cell(row=row_num, column=price_col).value
                if val is not None:
                    prices[section] = round(float(val))

            # Extract discount % (decimal -> integer %)
            disc_val = ws.cell(row=20, column=disc_col).value
            discount = 0
            if disc_val is not None:
                try:
                    discount = round(abs(float(disc_val)) * 100)  # abs() for negative values
                except (ValueError, TypeError):
                    discount = 0

            if prices:
                state_data[grade] = {**prices, "discount": discount}

        all_prices[state] = state_data

    wb.close()

    # Organize by grade for output
    result = {
        "effective_date": effective_date,
        "Fe 550": {},
        "Fe 550D": {},
    }

    for state in FE550_STATES:
        data = all_prices.get(state, {}).get("Fe 550")
        if data:
            result["Fe 550"][state] = data

    for state in FE550D_STATES:
        data = all_prices.get(state, {}).get("Fe 550D")
        if data:
            result["Fe 550D"][state] = data

    return result


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python tools/parse_ecp_prices.py <path_to_excel> <effective_date>", file=sys.stderr)
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    date = sys.argv[2]
    data = parse_excel(filepath, date)
    print(json.dumps(data, indent=2, ensure_ascii=False))
